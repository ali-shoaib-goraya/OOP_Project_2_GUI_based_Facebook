from tkinter import Button, Label, messagebox
from openpyxl import Workbook, load_workbook
from numpy import NaN, s_
from tkinter import*
# import GUI_Facebook
#wb = Workbook()         #Use this command to create workbook
#wb.save('Facebook.xlsx')   #Use this to save file O
wb = load_workbook('Facebook.xlsx')
wb.save('Facebook.xlsx')
class Facebook:
    wb = load_workbook('Facebook.xlsx')
    sh = wb['Sheet4']
    sh['A1'] = 'Friends_Posts'
    sh['A2'] = 'Emails'
    sh['B2'] = 'Posts'
    wb.save('Facebook.xlsx')

    wb = load_workbook('Facebook.xlsx')
    Sh = wb['Sheet']
    wb.save('Facebook.xlsx')
    Sh['A1'] = 'Email'
    Sh['B1'] = 'Password'
    Sh['C1'] = 'Name'
    Sh['D1'] = 'DOB'
    Sh['E1'] = 'Work_experience'
    Sh['F1'] = 'Bio'
    Sh['G1'] = 'Education'

    def __init__(self):
        self.name = None
        self.password = None
        self.email = None
        self.bio = None
        self.work_experience = None
        self.education = None
        self.Dob = None
        self.Logged_in = False

    def check_signup(self, email):
        self.emails= []
        end_row = Facebook.Sh.max_row+1
        for i in range(2,end_row):
            valu = Facebook.Sh.cell(i,1).value
            self.emails.append(valu)
        if email in self.emails:
            return True
        else:
            return False

################ To create account of that Roll_no which don't have account  ##################
    def Signup(self,email,__password,name,Dob): 
        sign_up = Facebook.check_signup(self, email)
        if sign_up== True:
            print('You already have an account! move to login')
            messagebox.showinfo("Signup",'You already have an account! move to login')
        else:
            self.email = email
            self.__password =__password
            self.Dob = Dob
            self.name = name
            lst = []
            lst.append(self.email)
            lst.append(self.__password)
            lst.append(self.name)
            lst.append(self.Dob)
            last_row = Facebook.Sh.max_row+1
            for i in range(len(lst)):
                Facebook.Sh.cell(last_row,column=i+1, value=lst[i])   # To add the value in cell
            Facebook.wb.save("Facebook.xlsx")
            Sh1 = Facebook.wb['Sheet1']
            Sh1['A1'] = 'Friend_requests_file'
            Facebook.wb.save('Facebook.xlsx')
            last_col = Sh1.max_column+1
            Sh1.cell(1,last_col,value=self.email)
            Facebook.wb.save('Facebook.xlsx')

            Sh2 = Facebook.wb['Sheet2']
            Sh2['A1'] = 'Friend_file'
            Facebook.wb.save('Facebook.xlsx')
            last_col = Sh2.max_column+1
            Sh2.cell(1,last_col,value=self.email)
            Facebook.wb.save('Facebook.xlsx')


            sh1 = Facebook.wb['Sheet5']
            sh1['A1'] = 'Notification_file'
            last = sh1.max_column+1
            sh1.cell(1,last,value=self.email)
            Facebook.wb.save('Facebook.xlsx')
            messagebox.showinfo("Signup","Your account is created successfully")


    def Login(self,email,name,password):
        a = False 
        sign_up = Facebook.check_signup(self,email)
        if sign_up == True:
            self.name_lst = []
            end_row = Facebook.Sh.max_row+1
            for i in range(1,end_row):
                valu = Facebook.Sh.cell(i,3).value
                self.name_lst.append(valu)
            if name in self.name_lst:
                index = self.name_lst.index(name)
                if name == self.name_lst[index]:
                    self.email_lst = []
                    end_row = Facebook.Sh.max_row+1
                    for i in range(1,end_row):
                        valu = Facebook.Sh.cell(i,1).value
                        self.email_lst.append(valu)
                    if email in self.email_lst:
                        index = self.email_lst.index(email)
                    if email == self.email_lst[index]:
                        self.password_lst = []
                        end_row = Facebook.Sh.max_row+1
                        for i in range(1,end_row):
                            p = Facebook.Sh.cell(i,2).value
                            self.password_lst.append(p) 
                        if password == self.password_lst[index]:
                            self.Logged_in = True
                            self.email = email
                            self.password = password
                            self.name = name
                            a = True  
                
                        else:
                            print('Invalid Password!   Try again') 
                            messagebox.showinfo("Login",'Invalid Password!   Try again') 
                                
                    else: 
                        print('Invalid Email!   Try again')         
                        messagebox.showinfo("Login",'Invalid Email!   Try again')
            else:
                print('Invalid Name!   Try again')  
                messagebox.showinfo("Login",'Invalid Name!   Try again')
        else:
                print("You don't have an account! move to sign_up")           
                messagebox.showinfo("Login","You don't have an account! move to sign_up")               
        return a

    def Return_column_values_lst(self,sh,s_row,e_row,col):
        lst = []
        for row in range(s_row,e_row):
            valu = sh.cell(row,col).value
            lst.append(valu)
        return(lst)

    def Return_Row_values_lst(self,sh,s_col,e_col,row):
        lst = []
        for col in range(s_col,e_col):
            valu = sh.cell(row,col).value
            lst.append(valu)
        return(lst)

    def Return_index_in_lst(self,lst,key):
        if key in lst:
            index = lst.index(key)
            return(index)



class User:

    def __init__(self):
        self.obj1 = Facebook()


    def Add_bio(self,bio):
        email = obj1.email
        obj1.bio = bio
        email_lst = []
        end_row = Facebook.Sh.max_row+1
        for i in range(1,end_row):
            valu = Facebook.Sh.cell(i,1).value
            email_lst.append(valu)
        if email in email_lst:
            index = email_lst.index(email)
            Facebook.Sh.cell(index+1, 6, value=obj1.bio)
            Facebook.wb.save('Facebook.xlsx')
            print('Your bio has been has been added')
            messagebox.showinfo("Bio","Bio has been updated successfully")
            

       
    def Add_Work(self,work):
        email = obj1.email 
        obj1.work_experience = work
        email_lst = []
        end_row = Facebook.Sh.max_row+1
        for i in range(1,end_row):
            valu = Facebook.Sh.cell(i,1).value
            email_lst.append(valu)
        if email in email_lst:
            index = email_lst.index(email)
            Facebook.Sh.cell(index+1, 5, value=obj1.work_experience)
            Facebook.wb.save('Facebook.xlsx')
            print('Your work details has been added')
            messagebox.showinfo("Bio","Your Work details has been updated successfully")
        
    def Add_Education(self,edu):
        email = obj1.email     
        obj1.education = edu
        email_lst = []
        end_row = Facebook.Sh.max_row+1
        for i in range(1,end_row):
            valu = Facebook.Sh.cell(i,1).value
            email_lst.append(valu)
        if email in email_lst:
            index = email_lst.index(email)
            Facebook.Sh.cell(index+1, 7, value=obj1.education)
            Facebook.wb.save('Facebook.xlsx')
            print('Your education details has been added')
            messagebox.showinfo("Bio","Your Education details has been updated successfully")

    
    def Search_member_by_name(self,aa,root):  
        name = aa
        wb = load_workbook("Facebook.xlsx")
        sh = wb['Sheet']
        names_lst = []
        last_row = sh.max_row+1
        for i in range(2,last_row):
            valu = sh.cell(i,3).value
            names_lst.append(valu)
        if name in names_lst:
            index = names_lst.index(name)
            row = index+2
            dd = str(name)+"'s  profile:"
            print(dd)
            lbl1 = Label(root,text=dd,fg='#00FF00',font=('Italic',20,'bold')).grid(row=2, column=0)
            nn = sh.cell(row, 3).value
            zz = 'Name:  '+nn
            print(zz)
            lbl2 = Label(root,text=zz,fg='#00FF00',font=('Italic',16,'bold')).grid(row=3, column=0)
            dob = sh.cell(row,4).value
            yy = 'Date of Birth:  '+dob
            print(yy)
            lbl3 = Label(root,text=yy,fg='#00FF00',font=('Italic',16,'bold')).grid(row=4, column=0)
            bio = sh.cell(row, 6).value
            if bio != None:
                ff = 'Bio:  '+bio
                print(ff)
                lbl4 = Label(root,text=ff,fg='#00FF00',font=('Italic',16,'bold')).grid(row=5, column=0)
            edu = sh.cell(row, 7).value
            if edu != None:
                gg = 'Education:  '+edu
                print(gg)
                lbl5 = Label(root,text=gg,fg='#00FF00',font=('Italic',16,'bold')).grid(row=6, column=0)
            work = sh.cell(row, 5).value
            if work != None:
                kk = 'Work Experience:  '+work
                print(kk)
                lbl6 = Label(root,text=kk,fg='#00FF00',font=('Italic',16,'bold')).grid(row=7, column=0)
        else:
            print('There is no member of this name!')
            messagebox.showinfo("Profile",'There is no member of this name!')



class Friends:

    def __init__(self):
        self.obj1 = Facebook()          ##   Compoosition  ##
        self.sent = False
        self.accept = False
        self.reject = False


    def get_max_row_in_col(self,ws, column):
        return max([cell[0] for cell in ws._cells if cell[1] == column])
    
    def Return_friends_lst(self):
        if obj1.Logged_in == True:
            user = obj1.email
            wb = load_workbook('Facebook.xlsx')
            sh = wb['Sheet2']
            last_col = sh.max_column+1
            lst = []
            for i in range(2,last_col):
                valu = sh.cell(1,i).value
                lst.append(valu)
            if user in lst:
                index = lst.index(user)
                col = index+2
                last_row = obj3.get_max_row_in_col(sh,col)+1
                friends = []
                for i in range(1,last_row):
                    valu = sh.cell(i,col).value
                    friends.append(valu)  
                return(friends)
            else:
                print(user, 'has no friends')
        else:
            print('You are not logged in!')

    def Return_Friend_requests_lst(self):
        user = obj1.email
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet1']
        last_col = sh.max_column+1
        lst = []
        for i in range(2,last_col):
            valu = sh.cell(1,i).value
            lst.append(valu)
        if user in lst:
            index = lst.index(user)
        col = index+2
        last_row = Friends.get_max_row_in_col(self,sh,col)+1
        f_lst = []
        for row in range(2,last_row):
            valu = sh.cell(row,col).value
            f_lst.append(valu)
        return(f_lst)


    def Return_Reciever_requests(self,reciever):
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet1']
        last_col = sh.max_column+1
        lst = []
        for i in range(2,last_col):
            valu = sh.cell(1,i).value
            lst.append(valu)
        if reciever in lst:
            index = lst.index(reciever)
        col = index+2
        last_row = Friends.get_max_row_in_col(self,sh,col)+1
        f_lst = []
        for row in range(2,last_row):
            valu = sh.cell(row,col).value
            f_lst.append(valu)
        return(f_lst)

    def Send_friend_request(self,ee):
        user = obj1.email   ## Use of composition ##
        nn = obj1.email
        choice = ee
        wb = load_workbook('Facebook.xlsx')
        sh1 = wb['Sheet']
        end = sh1.max_row+1
        Account_lst = obj1.Return_column_values_lst(sh1,1,end,1)
        if choice in Account_lst:
            friend_requests = Friends.Return_Friend_requests_lst(self)
            R_lst = Friends.Return_Reciever_requests(self,choice)
            if user not in R_lst:
                if choice not in friend_requests:
                    frind_lst = Friends.Return_friends_lst(self)
                    if choice not in frind_lst:
                        wb = load_workbook('Facebook.xlsx')
                        sh = wb['Sheet1']
                        col = sh.max_column+1
                        name_lst = []
                        for i in range(2,col):
                            val = sh.cell(1,i).value
                            name_lst.append(val)
                        if choice in name_lst:
                            index = name_lst.index(choice)
                            col = index+2
                            last_row = Friends.get_max_row_in_col(self,sh,col)
                            last_row+= 1
                            sh.cell(last_row,col,value=nn)
                            wb.save('Facebook.xlsx')
                            self.sent = True
                            print('Friend request has been sent')
                            messagebox.showinfo("Send Friend Request",'Friend request has been sent')
                            Noti = user+' has sent you a friend request'
                            sh2 = wb['Sheet5']
                            l_col = sh2.max_column+1
                            lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                            for i in range(len(lst)):
                                if choice in lst:
                                    index = lst.index(choice)+2
                                    last_row = obj3.get_max_row_in_col(sh2,index)+1
                                    sh2.cell(last_row,index,value=Noti)
                                    wb.save('Facebook.xlsx')
                                    break
                                else:
                                    print('user is not in header')
                                    messagebox.showinfo("Send Friend Request",'user is not in header')
                        else:
                            print('There is no such member!')
                            messagebox.showinfo("Send Friend Request",'There is no such member!')
                    else:
                        print('You are already friends!')
                        messagebox.showinfo("Send Friend Request",'You are already friends!')
                else:
                    print('You already have friend request of this person')
                    messagebox.showinfo("Send Friend Request",'You already have friend request of this person')
            else:
                print('You already have sent friend request!')
                messagebox.showinfo("Send Friend Request",'You already have sent friend request!')
        else:
            print('There is no such member!')
            messagebox.showinfo("Send Friend Request",'There is no such member!')
    
    def Accept_friend_request(self,aa):
        choice = aa
        user  = obj1.email
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet1']
        col1 = sh.max_column+1
        name_lst = []
        for i in range(2,col1):
            val = sh.cell(1,i).value
            name_lst.append(val)

        if user in name_lst:
            index = name_lst.index(user)
            col = index+2
            last_row = Friends.get_max_row_in_col(self,sh,col)
            last_row+= 1
            lst = []
            for i in range(2,last_row):
                valu = sh.cell(i,col).value
                lst.append(valu)
            if choice in lst:
                index = lst.index(choice)
                row = index+2
                sh.cell(row, col, value=NaN)
                wb = load_workbook('Facebook.xlsx')
                sh = wb['Sheet2']
                col1 = sh.max_column+1
                name_lst = []
                for i in range(2,col1):
                    val = sh.cell(1,i).value
                    name_lst.append(val)
                index = name_lst.index(user)
                index2 = name_lst.index(choice)
                col = index+2
                col2 = index2+2
                last_row = Friends.get_max_row_in_col(self,sh,col)
                last_row+= 1
                sh.cell(last_row,col,value=choice) 
                last_row2 = Friends.get_max_row_in_col(self,sh,col2)  ### to add in friend list of 
                last_row2+= 1
                sh.cell(last_row2,col2,value=user)
                wb.save('Facebook.xlsx')
                print("You and",choice,'are now friends')
                messagebox.showinfo("Accept Friend Request",'You and '+choice+' are now friends')

                Noti = user+' has accepted your friend request'
                sh2 = wb['Sheet5']
                l_col = sh2.max_column+1
                lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                if choice in lst:
                    index = lst.index(choice)+2
                    last_row = obj3.get_max_row_in_col(sh2,index)+1
                    sh2.cell(last_row,index,value=Noti)
                    wb.save('Facebook.xlsx')
                
                Sh = wb['Sheet1']
                col = Sh.max_column+1
                name_lst = []
                for i in range(2,col):
                    val = Sh.cell(1,i).value
                    name_lst.append(val)
                if user in name_lst:
                    index = name_lst.index(user)
                    col = index+2
                    last_row = Friends.get_max_row_in_col(self,Sh,col)
                    last_row+= 1
                    lst = []
                    for i in range(2,last_row):
                        valu = Sh.cell(i,col).value
                        lst.append(valu)
                    if choice in lst:
                        index = lst.index(choice)
                        row = index+2
                        Sh.cell(row, col, value = NaN)
                        wb.save('Facebook.xlsx')
            else:
                print('There is no friend request of', choice)
                messagebox.showinfo("Accept Friend Request",'There is no friend request of '+choice)
                

    def Reject_friend_request(self,aa):
        choice = aa
        user = obj1.email
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet1']
        col = sh.max_column+1
        name_lst = []
        for i in range(2,col):
            val = sh.cell(1,i).value
            name_lst.append(val)
        if user in name_lst:
            index = name_lst.index(user)
            col = index+2
            last_row = Friends.get_max_row_in_col(self,sh,col)
            last_row+= 1
            lst = []
            for i in range(2,last_row):
                valu = sh.cell(i,col).value
                lst.append(valu)
            if choice in lst:
                sh1 = wb ['Sheet5']
                cc = sh1.max_column
                lst3 =[]
                for i in range(2,cc):
                    valu = sh1.cell(1,i).value
                    lst3.append(valu)
                for j in range(len(lst3)):
                    count = 2
                    if choice == lst3[j]:
                        index = j
                        break
                    else:
                        count+=1
                col3 = index+count
                col3+=1
                last_row = obj3.get_max_row_in_col(sh1,col3)+1
                sh1.cell(last_row, col3,value=user)
                wb.save('Facebook.xlsx')
                index = lst.index(choice)
                row = index+2
                sh.cell(row, col, value = NaN)
                wb.save('Facebook.xlsx')
                print('Friend request of', choice,'is rejected')
                messagebox.showinfo("Reject Friend Request",'Friend request of '+choice+' is rejected')
                self.reject = True

            else:
                print('There is no friend request of',choice)
                messagebox.showinfo("Reject Friend Request",'There is no friend request of '+choice)
    def Print_Suggestions(self,root):
        if obj1.Logged_in == True:
            print('Following are the Suggestions:')
            Lbl = Label(root,text='Following are the Suggestions:',fg='#0000FF',font=('Italic',14,'bold')).grid(row=0,column=1)
            wb = load_workbook('Facebook.xlsx')
            sh = wb['Sheet']
            last_row = Friends.get_max_row_in_col(self,sh,1)+1
            lst = []
            for row in range(2,last_row):
                valu = sh.cell(row,1).value
                lst.append(valu)

            friend_lst = Friends.Return_friends_lst(self)
            cc = 1
            for j in range(len(lst)):
                if lst[j] not in friend_lst:
                    Lbl = Label(root,text=lst[j],font=(14))
                    Lbl.grid(row=cc,column=1)
                    print(lst[j])
                    cc+=1
    
    def Print_friend_requests(self,root):
        user = obj1.email
        print('Friend Requests are given below:')
        Lbl = Label(root,text='Following are the Friend requests:',fg='#0000FF',font=('Italic',14,'bold')).grid(row=0,column=1)
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet1']
        L_col = sh.max_column+1
        lst = []
        for i in range(2,L_col):
            valu = sh.cell(1,i).value
            lst.append(valu)
        if user in lst:
            index = lst.index(user)
        col = index+2
        last_row =Friends.get_max_row_in_col(self,sh,col)+1
        count = 1
        for k in range(2,last_row):
            valu = sh.cell(k,col).value
            if valu != None:
                Lbl = Label(root,text=valu,font=(14))
                Lbl.grid(row=count,column=1)
                count+= 1
                print(valu)



class Post:
    j = 0
    def __init__(self):
        self.obj1 = Facebook()          ##   Composition  ##
        self.obj3 = Friends()

    def Share_post_with_friends(self,cntnt):
        if obj1.Logged_in == True:
            if cntnt != '':
                pp = obj1.name+":"+cntnt
                friends_lst = obj3.Return_friends_lst()
                user = obj1.email
                wb = load_workbook('Facebook.xlsx')
                sh = wb['Sheet4']
                last_col = sh.max_column+1
                l_row = sh.max_row+1
                sh.cell(l_row,1,value=user)
                sh.cell(l_row,2,value=pp)
                wb.save('Facebook.xlsx')
                print('Post has been shared with friends.....')
                messagebox.showinfo('Share post','Post has been shared with friends.....')

                Noti = user+' has shared a post'
                sh2 = wb['Sheet5']
                l_col = sh2.max_column+1
                lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                for i in range(len(lst)):
                    for j in range(len(friends_lst)):
                        if friends_lst[j] in lst:
                            if friends_lst[j] != user:
                                index = lst.index(friends_lst[j])+2
                                last_row = obj3.get_max_row_in_col(sh2,index)+1
                                sh2.cell(last_row,index,value=Noti)
                                wb.save('Facebook.xlsx')
                    break
            else:
                print('Please write something to share post')
                messagebox.showinfo('Share post','Please write something to share post')
        else:
            print('You are not logged in!')


    def Display_post(self):
        root = Tk()
        root.geometry("700x500")
        self.root = root
        self.user = obj1.email
        friends_lst = obj3.Return_friends_lst()
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet4']
        last_row = sh.max_row+1
        email_lst = obj1.Return_column_values_lst(sh,3,last_row,1)
        post_lst = obj1.Return_column_values_lst(sh,3,last_row,2)
        if email_lst[Post.j] in friends_lst:
            if post_lst[Post.j] != None:
                print('Post is given below:')
                lbl1 = Label(root,text='Post is given below:',fg='#0000FF',font=('Calibari',18,'bold')).grid(row=0,column=0)
                print(post_lst[Post.j])
                lbl2 = Label(root,text=post_lst[Post.j],fg='#00FF00',font=('Calibari',14,'bold')).grid(row=1,column=0)
               
        Button(root,text='Enter your Comment',bg='White',fg='#0000FF',font=(14)).grid(row=4,column=0)
        bb = Entry(root)
        bb.grid(row=4,column=1)
        Button(root,text='Submit',fg='#0000FF',bg='White',font=('Calibari',16,'bold'),command=lambda:obj4.Add_Comment(bb.get())).grid(row=5,column=1)
        Button(root,text='See Comments',bg='White',fg='#0000FF',font=('Calibari',16,'bold'),command=lambda:obj4.See_Comments(root)).grid(row=6,column=1)
        Button(root,text='Next Post',bg='White',fg='#0000FF',font=('Calibari',16,'bold'),command=lambda:obj4.Next_post()).grid(row=6,column=2)
        Button(root,text='Previous Post',bg='White',fg='#0000FF',font=('Calibari',16,'bold'),command=lambda:obj4.Previous_post()).grid(row=6,column=0)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=a1.Quit).grid(row=6,column=4)
        root.mainloop()


    def Add_Comment(self,comment):
        sh = wb['Sheet4']
        S_col = 3
        row = Post.j+3
        comment = self.user+':'+comment
        while True :
            if sh.cell(row,S_col).value == None:
                sh.cell(row,S_col,value=comment)
                wb.save('Facebook.xlsx')
                print('Comment has been Added')
                messagebox.showinfo('Add Comment','Comment has been Added')
                break
            else:
                S_col+=1


    def See_Comments(self,root):
        sh = wb['Sheet4']
        S_col = 3
        row = Post.j+3
        print('Comments are given below:')
        c =0
        k = 7
        while True:
            if sh.cell(row,S_col).value != None:
                c+=1
                gg = sh.cell(row,S_col).value
                print(gg)
                Label(root,text=gg,font=(16)).grid(row=k,column=1)
                S_col+=1
                k+=1
                wb.save('Facebook.xlsx')
            else:
                break
        if c == 0:
            print('There are no Comments yet!')
            messagebox.showinfo('See Posts','There are no Comments yet!')
        
    def Next_post(self):
        Post.j+=1
        self.Display_post()

    def Previous_post(self):
        Post.j-=1
        self.Display_post()


    def Search_post_by_word(self,ch,root):
        choice = ch
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet4']
        last_row =sh.max_row+1
        lst =[]
        for i in range(3,last_row):
            valu = sh.cell(i,2).value
            lst.append(valu)
        cc=3
        if lst != []:
            for i in range(len(lst)):
                if lst[i] != None:
                    if choice in lst[i]:
                        lbl = Label(root,text=lst[i],font=(16))
                        lbl.grid(row=cc,column=0)
                        cc+=1
                        # self.search_again(lbl,root)
                        print(lst[i])
        else:
            print('There is no post yet')
            messagebox.showinfo("Search post by word",'There is no post yet')

    def See_Notifications(self,root):
        wb = load_workbook("Facebook.xlsx")
        user = obj1.email
        sh = wb['Sheet5']
        e_col = sh.max_column+1
        header = obj1.Return_Row_values_lst(sh,2,e_col,1)
        if user in header:
            index = header.index(user)+2
        end_row = obj3.get_max_row_in_col(sh,index)+1
        lst = obj1.Return_column_values_lst(sh,2,end_row,index)
        c = len(lst)-1
        if len(lst)!= 0:
            print('Notifications are given below:')
            lbl1 = Label(root, text='Notifications are given below:',fg='#0000FF',font=('Italic',20,'bold')).place(x=20,y=1)
            count = 2
            for n in range(len(lst)):
                lbl2 = Label(root, text=lst[c],fg='#00FF00',font=(16)).place(x=50,y=count*25)
                print(lst[c])
                c = c-1
                count+=1
        else:
            print('There are no notifications')
            messagebox.showinfo('Notification','There is no any notification')


class Message:

    def __init__(self):
        self.obj1 = Facebook()          ##   Composition  ##
        self.obj3 = Friends()


    def Sent_message(self,receiver,message):
        wb = load_workbook('Facebook.xlsx')
        sh1 = wb['Sheet']
        user = obj1.email
        Accounts_lst = []
        row = sh1.max_row+1
        for i in range(2,row):
            valu = sh1.cell(i,1).value
            Accounts_lst.append(valu)
        if receiver in Accounts_lst:
            if receiver != user:
                sh = wb['Sheet6']
                message = user+':'+message
                head = user+' and '+receiver
                head2 = receiver+' and '+user

                sh['A1'] = 'Messages_file'
                last_col = sh.max_column+1
                head_lst = []
                for k in range(2,last_col):
                    valu = sh.cell(1,k).value
                    head_lst.append(valu)
                if head not in head_lst and head2 not in head_lst:
                    sh.cell(1,last_col,value=head)
                    last = sh.max_row+1
                    sh.cell(last,last_col,value=message)
                    wb.save('Facebook.xlsx')
                    print('message has been sent')
                    messagebox.showinfo('Sent message','message has been sent')

                    Noti = user+' has sent you a message'
                    sh2 = wb['Sheet5']
                    l_col = sh2.max_column+1
                    lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                    if receiver in lst:
                        index = lst.index(receiver)+2
                        last_row = obj3.get_max_row_in_col(sh2,index)+1
                        sh2.cell(last_row,index,value=Noti)
                        wb.save('Facebook.xlsx')

                elif head in head_lst:
                    index = head_lst.index(head)
                    f_col = index+2
                    last = sh.max_row+1
                    sh.cell(last,f_col,value=message)
                    wb.save('Facebook.xlsx')
                    print('message has been sent')
                    messagebox.showinfo('Sent message','message has been sent')

                    Noti = user+' has sent you a message'
                    sh2 = wb['Sheet5']
                    l_col = sh2.max_column+1
                    lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                    if receiver in lst:
                        index = lst.index(receiver)+2
                        last_row = obj3.get_max_row_in_col(sh2,index)+1
                        sh2.cell(last_row,index,value=Noti)
                        wb.save('Facebook.xlsx')
                elif head2 in head_lst:
                    index = head_lst.index(head2)
                    f_col = index+2
                    last = sh.max_row+1
                    sh.cell(last,f_col,value=message)
                    wb.save('Facebook.xlsx')
                    print('message has been sent')
                    messagebox.showinfo('Sent message','message has been sent')
                    Noti = user+' has sent you a message'
                    sh2 = wb['Sheet5']
                    l_col = sh2.max_column+1
                    lst = obj1.Return_Row_values_lst(sh2, 2, l_col,1)
                    if receiver in lst:
                        index = lst.index(receiver)+2
                        last_row = obj3.get_max_row_in_col(sh2,index)+1
                        sh2.cell(last_row,index,value=Noti)
                        wb.save('Facebook.xlsx')

            else:
                print("You can't send message to yourself!")
                messagebox.showinfo('Sent message',"You can't send message to yourself!")
        else:
            print('There is no such member!')
            messagebox.showinfo('Sent message','There is no such member!')

    def Print_message_Suggestions(self,root):
        user = obj1.email
        print('Following are the Suggestions:')
        Label(root,text='Following are the Suggestions:',fg='#0000FF',font=('Italian',18,'bold')).grid(row=0,column=0)
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet']
        last_row = obj3.get_max_row_in_col(sh,1)+1
        lst = []
        for row in range(2,last_row):
            valu = sh.cell(row,1).value
            lst.append(valu)
        cc = 1
        for j in range(len(lst)):
            if lst[j] != user:
                print(lst[j])
                Label(root,text=lst[j],font=(15)).grid(row=cc,column=0)
                cc+=1


    def Display_messages(self,root,other):
        user = obj1.email
        wb = load_workbook('Facebook.xlsx')
        sh = wb['Sheet6']
        ss = user+' and '+other
        ss2 = other+' and '+user
        last = sh.max_column+1
        lst = []
        for i in range(2,last):
            valu = sh.cell(1,i).value
            lst.append(valu)
        if ss in lst:
            Label(root,text='Chat is given below:').grid(row=4,column=8)
            index = lst.index(ss)
            col = index+2
            last_row = obj3.get_max_row_in_col(sh,col)+1
            k = 5
            for r in range(2,last_row):
                valu = sh.cell(r,col).value
                print(valu)
                Label(root,text=valu).grid(row=k,column=8)
                k+=1
        elif ss2 in lst:
            Label(root,text='Chat is given below:').grid(row=4,column=8)
            index = lst.index(ss2)
            col = index+2
            last_row = obj3.get_max_row_in_col(sh,col)+1
            k = 5
            for r in range(2,last_row):
                valu = sh.cell(r,col).value
                print(valu)
                Label(root,text=valu).grid(row=k,column=8)
                k+=1
        else:
            print('There is no chat in between you and',other)
            messagebox.showinfo('See Chat','There is no chat in between you and '+other)


class Page:
    pass

class Privacy:
    pass


print('....................Welcome to Console based Facebook...................')
obj1= Facebook() 
obj2 = User()
obj3 = Friends()
obj4 = Post()
obj5 = Message()

#################################################################################################################
class GUI:
    def show_initial_Page(self):
        root = Tk()
        #root.attributes("-fullscreen",True)
        root.geometry('700x500')
        Label(root,text = "Welcome to facebook",fg='#0000FF',bg='White',font=('Impact',35,'bold')).grid(row = 1,column = 5)
        Label(root,text = "Already have a account",font=('Impact',20)).grid(row = 4,column = 0)
        Button(root,text = "login",fg='#0000FF',bg='White',font=('Impact',20,'bold'),command =a1.login).grid(row =  5,column = 0)
        Label(root,text = "To create new account",font=('Impact',20)).grid(row = 7,column = 0)
        Button(root,text = "signup",fg='#0000FF',bg='White',font=('Impact',20,'bold'),command =a1.signup).grid(row =  8,column = 0)
        root.mainloop()



    def login(self):
        root = Tk()
        root.geometry("700x500")
        Label(root,text = "To login verify yourself ",fg='#0000FF',font=('Calibari',20,'bold')).grid(row = 0,column =  0)
        Label(root,text = "username",font=(20)).grid(row = 2,column = 0)
        name = Entry(root)            
        name.grid(row = 2,column = 1)
        Label(root,text = "password",font=(20)).grid(row = 3,column = 0)
        password = Entry(root)
        password.grid(row = 3,column = 1)
        Label(root,text = "email",font=(20)).grid(row = 4,column = 0)
        email = Entry(root) 
        email.grid(row = 4,column = 1)
        Button(root,text = "Verify",fg='#0000FF',font=('Calibary',20,'bold'),command = lambda:a1.check_login(email.get(),name.get(),password.get())).grid(row = 6,column = 0)
        root.mainloop()
    def check_login(self,email,name,password):
        y = obj1.Login(email,name,password)
        if y == True:
            messagebox.showinfo("loggedin","successfully logged in")
        self.show_main_page()
        


    def signup(self):
        root  =Tk()
        root.geometry("700x500")
        Label(root,text = "To create Account Enter your information ",fg='#0000FF',font=('Calibari',20,'bold')).grid(row = 0,column =  0)
        Label(root,text = "username",font=(20)).grid(row = 2,column = 0)
        name = Entry(root)            
        name.grid(row = 2,column = 1)
        Label(root,text = "password",font=(20)).grid(row = 3,column = 0)
        password = Entry(root)
        password.grid(row = 3,column = 1)
        Label(root,text = "email",font=(20)).grid(row = 4,column = 0)
        email = Entry(root)
        email.grid(row = 4,column = 1)

        Label(root,text = "Date of birth",font=(20)).grid(row = 5,column = 0)
        dob = Entry(root)
        dob.grid(row = 5,column = 1)
        #print(email.get(),name.get(),password.get())
        Button(root,text = "Submit",fg='#0000FF',font=('Calibary',20,'bold'),command = lambda:obj1.Signup(email.get(),password.get(),name.get(),dob.get())).grid(row = 6,column = 0)
        root.mainloop()



    def show_main_page(self):
        root = Tk()
        #root.attributes("-fullscreen",True)
        root.geometry('700x500')
        Label(root,text = "Which function you want to perform click on that button",fg='#0000FF',font=('Calibari',20,'bold')).grid(row = 0,column = 0)
        Button(root,text = "Add Bio",fg='#0000FF',font=(20),command=a1.Bio).place(x=60,y=50)
        Button(root,text = "Add Education",fg='#0000FF',font=(20),command =a1.Education).place(x=150,y=50)
        Button(root,text = "Add work Experience",fg='#0000FF',font=(20),command =a1.Work).place(x=270,y=50)
        Button(root,text = "Sent friend Request",fg='#0000FF',font=(20),command =a1.Sent_Request).place(x=450,y=50)
        Button(root,text = "Accept friend Request",fg='#0000FF',font=(20),command =a1.Accept_Request).place(x=620,y=50)
        Button(root,text = "Reject friend Request",fg='#0000FF',font=(20),command =a1.Reject_Request).place(x=810,y=50)
        Button(root,text = "Search profile by name",fg='#0000FF',font=(20),command =a1.Search_profile_by_name).place(x=990,y=50)
        Button(root,text = "See Notifications",fg='#0000FF',font=(20),command =a1.See_Notifications).place(x=60,y=100)
        Button(root,text = "Share post with friends",fg='#0000FF',font=(20),command =a1.Share_post).place(x=200,y=100)
        Button(root,text = "Search post by word",fg='#0000FF',font=(20),command =a1.Search_post).place(x=400,y=100)
        Button(root,text = "See Posts",fg='#0000FF',font=(20),command =obj4.Display_post).place(x=580,y=100)
        Button(root,text = "Sent messages",fg='#0000FF',font=(20),command =a1.Sent_messages).place(x=690,y=100)
        Button(root,text = "See messages",fg='#0000FF',font=(20),command =a1.See_messages).place(x=840,y=100)
        Button(root,text='Logout',fg='#FF0000',font=('Arial',20,'bold'),command=self.Logout).place(x=100,y=150)
        root.mainloop()
    def Logout(self):
        self.login()

    def Quit(self):
        self.show_main_page()
    def Bio(self):
        root = Tk()
        root.geometry("700x500")
        Label(root,text = "Add Bio",font=('Calibari',20,'bold')).grid(row = 1,column = 0)
        name = Entry(root)            
        name.grid(row = 1,column = 1)
        Button(root,text='Submit',fg='#0000FF',font=(20),command=lambda:obj2.Add_bio(name.get())).grid(row=2,column=0)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=3,column=1)
        root.mainloop()
    def Education(self):
        root = Tk()
        root.geometry("700x500")
        Label(root,text = "Add Education",font=('Calibari',20,'bold')).grid(row = 1,column = 0)
        name = Entry(root)            
        name.grid(row = 1,column = 1)
        Button(root,text='Submit',fg='#0000FF',font=(20),command=lambda:obj2.Add_Education(name.get())).grid(row=2,column=0)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=3,column=1)
        root.mainloop()
    def Work(self):
        root = Tk()
        root.geometry("700x500")
        Label(root,text = "Add Work Details",font=('Calibari',20,'bold')).grid(row = 1,column = 0)
        name = Entry(root)            
        name.grid(row = 1,column = 1)
        Button(root,text='Submit',fg='#0000FF',font=(20),command=lambda:obj2.Add_Work(name.get())).grid(row=2,column=0)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=3,column=1)
        root.mainloop()
    def Sent_Request(self):
        root = Tk()
        root.geometry("700x500")
        obj3.Print_Suggestions(root)
        Label(root,text = "Enter the Email to whom you want to send friend request",fg='#0000FF',font=(14)).grid(row = 1,column = 8)
        name = Entry(root)        
        name.grid(row = 1,column = 9)
        Button(root,text='Submit',fg='#0000FF',font=(20),command=lambda:obj3.Send_friend_request(name.get())).grid(row=2,column=9)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=4,column=9)
        root.mainloop()
    def Accept_Request(self):
        root = Tk()
        root.geometry("700x500")
        obj3.Print_friend_requests(root)
        Label(root,text = "Enter the Email of which friend request you want to Accept",fg='#0000FF',font=(14)).grid(row = 1,column = 8)
        name = Entry(root)        
        name.grid(row = 1,column = 9)
        Button(root,text='Submit',fg='#0000FF',font=(14),command=lambda:obj3.Accept_friend_request(name.get())).grid(row=2,column=9)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=4,column=9)
        root.mainloop
    def Reject_Request(self):
        root = Tk()
        root.geometry("700x500")
        obj3.Print_friend_requests(root)
        Label(root,text = "Enter the Email of which friend request you want to Reject",fg='#0000FF',font=(14)).grid(row = 1,column = 8)
        name = Entry(root)            
        name.grid(row = 1,column = 9)
        Button(root,text='Submit',fg='#0000FF',font=(14),command=lambda:obj3.Reject_friend_request(name.get())).grid(row=2,column=9)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=4,column=9)
        root.mainloop()
    def Search_profile_by_name(self):
        root = Tk()
        root.geometry("700x500")
        Label(root,text = "Enter the Name of which profile you want to see",fg='#0000FF',font=('Italic',18,'bold')).grid(row = 1,column = 0)
        name = Entry(root)            
        name.grid(row = 1,column = 1)
        Button(root,text='Submit',fg='#0000FF',font=(14),command=lambda:obj2.Search_member_by_name(name.get(),root)).grid(row=2,column=1)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=3,column=1)
        root.mainloop()
    def See_Notifications(self):
        root = Tk()
        root.geometry("700x500")
        obj4.See_Notifications(root)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).place(x=400,y=100)
        root.mainloop()
    def Search_post(self):
        root = Tk()
        root.geometry("700x500")
        Label(root,text = "Enter the word to search post",fg='#0000FF',font=('Italian',20,'bold')).grid(row = 0,column = 0)
        name = Entry(root)
        name.grid(row=0,column=1)
        Button(root,text='Search',fg='#0000FF',font=(14),command=lambda:obj4.Search_post_by_word(name.get(),root)).grid(row=1,column=1)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=2,column=1)
    def Share_post(self):
        root = Tk()
        root.geometry("700x500")
        Label(root,text = "Write your post",fg='#0000FF',font=('Italian',20,'bold')).grid(row = 0,column = 0)
        name = Entry(root)
        name.grid(row=0,column=1)
        Button(root,text='Share',fg='#0000FF',font=(14),command=lambda:obj4.Share_post_with_friends(name.get())).grid(row=1,column=1)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=3,column=1)
    def Sent_messages(self):
        root = Tk()               
        root.geometry("700x500")             
        obj5.Print_message_Suggestions(root)                
        Label(root,text = "Enter the Email to whom you want to send message",fg='#0000FF',font=(14)).grid(row = 1,column = 8)
        name = Entry(root)            
        name.grid(row = 1,column = 9)
        Label(root,text='Enter your message',fg='#0000FF',font=(14)).grid(row=2,column=8)
        name2 = Entry(root)
        name2.grid(row=2,column=9)
        Button(root,text='Send message',fg='#0000FF',font=(16),command=lambda:obj5.Sent_message(name.get(),name2.get())).grid(row=3,column=9)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=4,column=9)
        root.mainloop()
        
    def See_messages(self):
        root = Tk()
        root.geometry("700x500")
        obj5.Print_message_Suggestions(root)
        Label(root,text = 'Enter the email of person whoose messages you want to see:',fg='#0000FF',font=(14)).grid(row = 1,column = 8)
        name = Entry(root)            
        name.grid(row = 1,column = 9)
        Button(root,text='Submit ',fg="#0000FF",font=(16),command=lambda:obj5.Display_messages(root,name.get())).grid(row=2,column=9)
        Button(root,text='Quit',fg='#FF0000',font=('Arial',20,'bold'),command=self.Quit).grid(row=4,column=9)
        root.mainloop()    

   
a1 = GUI()
a1.show_initial_Page()


